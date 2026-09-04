"""Professional, generation-only XLSX adapter for the fixed client template."""

from __future__ import annotations

import hashlib
import io
import math
import os
import re
import zipfile
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from .client_export_dataset import (
    ClientExportDatasetError,
    _canonical_digest,
    validate_client_export_dataset,
)
from .export_templates import ExportTemplateError, preflight_client_layout
from .image_evidence import inspect_image_bytes
from .safe_io import read_regular_bytes

RECEIPT_SCHEMA = "xlsx-render-receipt/v1"
MAX_OUTPUT_BYTES = 512 * 1024 * 1024
EXCEL_CELL_LIMIT = 32_767
EXCEL_MAX_ROWS = 1_048_576
XLSX_MAX_LINES_PER_ROW = 24
SHEET_NAMES = (
    "01_项目概览",
    "02_逐镜分镜表",
    "03_VO与画面文字",
    "04_音乐与节奏",
    "05_证据与说明",
)
EVENT_KINDS = ("voice", "music", "sfx", "silence", "mixed")
ILLEGAL_XML_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
CJK_RE = re.compile(r"[\u2e80-\u9fff\uf900-\ufaff]")


class XlsxExportError(ValueError):
    pass


def render_client_xlsx(
    dataset: dict[str, Any],
    output_path: Path,
    *,
    settings: dict[str, Any] | None = None,
    project_root: Path,
    available_fonts: list[str] | tuple[str, ...] = (),
) -> dict[str, Any]:
    """Render one new XLSX file from a validated, generation-bound dataset.

    The renderer never replaces an existing file. The export transaction owns
    staging, atomic publication, idempotency and artifact-registry state.
    """

    output = Path(output_path)
    root = Path(project_root).resolve()
    if output.suffix.lower() != ".xlsx":
        raise XlsxExportError("XLSX output path must end in .xlsx")
    if output.exists() or output.is_symlink():
        raise XlsxExportError("XLSX output already exists")
    if not output.parent.is_dir():
        raise XlsxExportError("XLSX output parent directory does not exist")
    if not root.is_dir():
        raise XlsxExportError("project root does not exist")

    api = _openpyxl_api()
    try:
        dataset = validate_client_export_dataset(dataset)
    except ClientExportDatasetError as exc:
        raise XlsxExportError(str(exc)) from exc
    renderer_settings = dict(settings or {})
    requested_formats = renderer_settings.get("formats")
    if requested_formats is not None and (
        type(requested_formats) is not list
        or any(type(item) is not str for item in requested_formats)
        or "xlsx" not in requested_formats
    ):
        raise XlsxExportError("XLSX renderer settings must include xlsx")
    renderer_settings["formats"] = ["xlsx"]
    try:
        plan = preflight_client_layout(
            dataset,
            renderer_settings,
            project_root=root,
            available_fonts=available_fonts,
        )
    except ExportTemplateError as exc:
        raise XlsxExportError(str(exc)) from exc

    rendered_at = datetime.now(UTC).replace(microsecond=0)
    workbook, stats, image_streams = _build_workbook(dataset, plan, root, api, rendered_at)
    created = False
    try:
        flags = os.O_WRONLY | os.O_CREAT | os.O_EXCL
        if hasattr(os, "O_NOFOLLOW"):
            flags |= os.O_NOFOLLOW
        descriptor = os.open(output, flags, 0o600)
        created = True
        with os.fdopen(descriptor, "wb") as handle:
            workbook.save(handle)
            handle.flush()
            os.fsync(handle.fileno())
        archive = _validate_xlsx_archive(output, expected_images=stats["embedded_image_count"])
        raw = read_regular_bytes(output, root=output.parent, max_bytes=MAX_OUTPUT_BYTES)
    except Exception as exc:
        if created:
            output.unlink(missing_ok=True)
        if isinstance(exc, XlsxExportError):
            raise
        raise XlsxExportError(f"XLSX rendering failed ({type(exc).__name__})") from exc
    finally:
        image_streams.clear()
        workbook.close()

    base = {
        "schema_id": RECEIPT_SCHEMA,
        "renderer": {"name": "openpyxl", "version": api["version"]},
        "template_id": plan["template"]["template_id"],
        "template_version": plan["template"]["template_version"],
        "template_digest": plan["template"]["template_digest"],
        "preflight_digest": plan["preflight_digest"],
        "dataset_id": dataset["dataset_id"],
        "dataset_digest": dataset["dataset_digest"],
        "settings": plan["settings"],
        "font_plan": plan["font_plan"],
        "generated_at_utc": rendered_at.isoformat(),
        "sha256": hashlib.sha256(raw).hexdigest(),
        "size_bytes": len(raw),
        "sheet_names": list(SHEET_NAMES),
        "sheet_count": archive["sheet_count"],
        "formula_count": archive["formula_count"],
        "external_link_count": archive["external_link_count"],
        "contains_macros": archive["contains_macros"],
        **stats,
    }
    return {**base, "receipt_digest": _canonical_digest(base)}


def _openpyxl_api() -> dict[str, Any]:
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.chart import Reference, ScatterChart, Series
        from openpyxl.drawing.image import Image
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.page import PageMargins
    except ImportError as exc:
        raise XlsxExportError(
            "XLSX export extra is not installed; install video-analysis-mvp[export]"
        ) from exc
    return {
        "version": openpyxl.__version__,
        "Workbook": Workbook,
        "ScatterChart": ScatterChart,
        "Series": Series,
        "Reference": Reference,
        "Image": Image,
        "Alignment": Alignment,
        "Border": Border,
        "Font": Font,
        "PatternFill": PatternFill,
        "Side": Side,
        "PageMargins": PageMargins,
    }


def _build_workbook(
    dataset: dict[str, Any],
    plan: dict[str, Any],
    project_root: Path,
    api: dict[str, Any],
    rendered_at: datetime,
) -> tuple[Any, dict[str, Any], list[io.BytesIO]]:
    workbook = api["Workbook"]()
    workbook.remove(workbook.active)
    for name in SHEET_NAMES:
        workbook.create_sheet(name)
    workbook.active = 0
    workbook.properties.title = _plain(dataset["project"]["title"])
    workbook.properties.subject = "Evidence-bound storyboard, VO, music, SFX and rhythm breakdown"
    workbook.properties.creator = "Video Evidence Workbench"
    workbook.properties.lastModifiedBy = "Video Evidence Workbench"
    workbook.properties.category = "Client evidence report"
    workbook.properties.keywords = "video evidence, storyboard, VO, music, SFX, rhythm"
    workbook.properties.language = plan["settings"]["language"]
    ooxml_time = rendered_at.astimezone(UTC).replace(tzinfo=None)
    workbook.properties.created = ooxml_time
    workbook.properties.modified = ooxml_time

    styles = _styles(plan, api)
    image_streams: list[io.BytesIO] = []
    overview_stats = _overview_sheet(workbook[SHEET_NAMES[0]], dataset, plan, styles, project_root, api, image_streams)
    storyboard_stats = _storyboard_sheet(workbook[SHEET_NAMES[1]], dataset, plan, styles, project_root, api, image_streams)
    voice_rows, voice_continuations = _voice_sheet(workbook[SHEET_NAMES[2]], dataset, plan, styles, api)
    (
        audio_rows,
        audio_visualization,
        audio_source_points,
        audio_points,
        audio_method,
        audio_continuations,
    ) = _audio_sheet(
        workbook[SHEET_NAMES[3]], dataset, plan, styles, api
    )
    evidence_rows, evidence_continuations = _evidence_sheet(workbook[SHEET_NAMES[4]], dataset, plan, styles, api)
    auxiliary_continuations = (
        overview_stats.pop("overview_continuation_row_count")
        + voice_continuations
        + audio_continuations
        + evidence_continuations
    )
    storyboard_continuations = storyboard_stats.pop("storyboard_continuation_row_count")
    embedded_images = int(overview_stats["logo_embedded"]) + storyboard_stats["embedded_frame_count"]
    return workbook, {
        "declared_font_name": styles["font_name"],
        "declared_font_verified": styles["font_verified"],
        "warnings": [*plan["warnings"], *styles["renderer_warnings"]],
        **overview_stats,
        **storyboard_stats,
        "embedded_image_count": embedded_images,
        "storyboard_continuation_row_count": storyboard_continuations,
        "auxiliary_continuation_row_count": auxiliary_continuations,
        "continuation_row_count": storyboard_continuations + auxiliary_continuations,
        "voice_text_row_count": voice_rows,
        "audio_event_row_count": audio_rows,
        "audio_visualization_present": audio_visualization,
        "audio_visualization_source_point_count": audio_source_points,
        "audio_visualization_point_count": audio_points,
        "audio_visualization_method": audio_method,
        "evidence_row_count": evidence_rows,
    }, image_streams


def _styles(plan: dict[str, Any], api: dict[str, Any]) -> dict[str, Any]:
    tokens = plan["settings"]
    font_name = plan["typography"]["font_stack"][0] if plan["font_plan"]["cjk_required"] else "Arial"
    font_verified = font_name in plan["font_plan"]["available_fonts"]
    renderer_warnings = []
    if plan["font_plan"]["cjk_required"] and not font_verified:
        renderer_warnings.append(_label(
            tokens["language"],
            f"工作簿声明字体 {font_name} 未在目标表格渲染器中验证；字体替换可能影响版式。",
            f"Declared workbook font {font_name} was not verified for the target spreadsheet renderer; substitution may affect layout.",
        ))
    colors = {
        "ink": "111827",
        "navy": "0B1020",
        "accent": tokens["accent_color"].removeprefix("#"),
        "teal": "237A70",
        "paper": "F7F9FC",
        "line": "D6DDEA",
        "muted": "667085",
        "warning": "8A5600",
        "white": "FFFFFF",
    }
    Font = api["Font"]
    PatternFill = api["PatternFill"]
    Alignment = api["Alignment"]
    Border = api["Border"]
    Side = api["Side"]
    thin = Side(style="thin", color=colors["line"])
    medium = Side(style="medium", color=colors["navy"])
    return {
        "font_name": font_name,
        "font_verified": font_verified,
        "renderer_warnings": renderer_warnings,
        "colors": colors,
        "title_font": Font(name=font_name, size=22, bold=True, color=colors["white"]),
        "subtitle_font": Font(name=font_name, size=9, color=colors["white"]),
        "header_font": Font(name=font_name, size=9.5, bold=True, color=colors["white"]),
        "body_font": Font(name=font_name, size=9.5, color=colors["ink"]),
        "meta_font": Font(name=font_name, size=8.5, color=colors["muted"]),
        "warning_font": Font(name=font_name, size=9, bold=True, color=colors["warning"]),
        "title_fill": PatternFill("solid", fgColor=colors["navy"]),
        "header_fill": PatternFill("solid", fgColor=colors["teal"]),
        "accent_fill": PatternFill("solid", fgColor=colors["accent"]),
        "paper_fill": PatternFill("solid", fgColor=colors["paper"]),
        "warning_fill": PatternFill("solid", fgColor="FFF4E5"),
        "body_alignment": Alignment(vertical="top", wrap_text=True),
        "center_alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "title_alignment": Alignment(horizontal="left", vertical="center"),
        "thin_bottom": Border(bottom=thin),
        "section_border": Border(bottom=medium),
    }


def _overview_sheet(
    sheet: Any,
    dataset: dict[str, Any],
    plan: dict[str, Any],
    styles: dict[str, Any],
    project_root: Path,
    api: dict[str, Any],
    image_streams: list[io.BytesIO],
) -> dict[str, Any]:
    language = plan["settings"]["language"]
    project = dataset["project"]
    delivery = dataset["delivery_status"]
    sheet.merge_cells("A1:H1")
    sheet["A1"] = _label(language, "客户视频拆解报告", "CLIENT VIDEO BREAKDOWN")
    sheet["A1"].font = styles["title_font"]
    sheet["A1"].fill = styles["title_fill"]
    sheet["A1"].alignment = styles["title_alignment"]
    for cell in sheet[1][1:8]:
        cell.fill = styles["title_fill"]
    sheet.row_dimensions[1].height = 34

    subtitle = _plain(plan["settings"]["project_subtitle"])
    status = _label(language, "专业可交付" if delivery["professional_export_allowed"] else "仅供草稿审核", "PROFESSIONAL" if delivery["professional_export_allowed"] else "DRAFT ONLY")
    subtitle_text = " · ".join(
        item for item in (subtitle, status, f"template {plan['template']['template_version']}") if item
    )
    subtitle_chunks = _split_text(subtitle_text)
    for subtitle_row, chunk in enumerate(subtitle_chunks, 2):
        sheet.merge_cells(start_row=subtitle_row, start_column=1, end_row=subtitle_row, end_column=8)
        sheet.cell(subtitle_row, 1, _safe_string(chunk))
        sheet.cell(subtitle_row, 1).font = styles["subtitle_font"]
        sheet.cell(subtitle_row, 1).fill = styles["title_fill"]
        sheet.cell(subtitle_row, 1).alignment = styles["title_alignment"]
        for cell in sheet[subtitle_row][1:8]:
            cell.fill = styles["title_fill"]
        sheet.row_dimensions[subtitle_row].height = _row_height([chunk], minimum=24)
    overview_continuations = len(subtitle_chunks) - 1

    metrics = [
        (_label(language, "项目", "Project"), _plain(project["title"]), _label(language, "状态", "Status"), status),
        (_label(language, "时长", "Duration"), f"{project['duration_seconds']:.2f} s", _label(language, "镜头", "Shots"), len(dataset["shots"])),
        (_label(language, "画面", "Frame"), f"{project['resolution']} · {project['frame_rate']:.3f} fps", _label(language, "段落", "Sections"), len(dataset["scenes"])),
        (_label(language, "画面比", "Aspect"), f"{project['aspect_ratio']:.3f}:1", _label(language, "音频事件", "Audio events"), len(dataset["audio"]["events"])),
        (_label(language, "分析模式", "Profile"), project["analysis_profile"], _label(language, "语言", "Language"), plan["settings"]["language"]),
        (_label(language, "数据绑定", "Dataset binding"), dataset["dataset_digest"], _label(language, "就绪状态", "Readiness"), delivery["readiness_status"]),
    ]
    row = 3 + len(subtitle_chunks)
    metric_start = row
    for left_label, left_value, right_label, right_value in metrics:
        sheet.cell(row, 1, _safe_string(left_label))
        sheet.cell(row, 2, _safe_scalar(left_value))
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        sheet.cell(row, 5, _safe_string(right_label))
        sheet.cell(row, 6, _safe_scalar(right_value))
        sheet.merge_cells(start_row=row, start_column=6, end_row=row, end_column=8)
        for column in (1, 5):
            sheet.cell(row, column).font = styles["header_font"]
            sheet.cell(row, column).fill = styles["header_fill"]
            sheet.cell(row, column).alignment = styles["center_alignment"]
        for column in (2, 3, 6, 7, 8):
            sheet.cell(row, column).font = styles["body_font"]
            sheet.cell(row, column).alignment = styles["body_alignment"]
            sheet.cell(row, column).fill = styles["paper_fill"]
        sheet.row_dimensions[row].height = 28
        row += 1

    row += 1
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    sheet.cell(row, 1, _label(language, "叙事段落", "NARRATIVE SECTIONS"))
    _section_cell(sheet.cell(row, 1), styles)
    row += 1
    scene_header = [_label(language, "段落", "Section"), _label(language, "时间", "Range"), _label(language, "功能", "Function"), _label(language, "节奏", "Pace"), _label(language, "镜头", "Shots"), _label(language, "置信度", "Confidence")]
    for column, value in enumerate(scene_header, 1):
        sheet.cell(row, column, _safe_string(value))
        _header_cell(sheet.cell(row, column), styles)
    scene_rows = dataset["scenes"] or [{"scene_id": "—", "start_seconds": 0.0, "end_seconds": project["duration_seconds"], "function": {"spreadsheet_text": "No reviewed narrative sections", "text": "No reviewed narrative sections", "is_blank": False, "formula_neutralized": False}, "pace": {"spreadsheet_text": "unknown", "text": "unknown", "is_blank": False, "formula_neutralized": False}, "shot_ids": [], "confidence": 0.0}]
    scene_values = [
        [
            scene["scene_id"],
            _time_range(scene["start_seconds"], scene["end_seconds"]),
            _plain(scene["function"]),
            _plain(scene["pace"]),
            len(scene["shot_ids"]),
            scene["confidence"],
        ]
        for scene in scene_rows
    ]
    scene_start = row + 1
    next_row, scene_continuations = _write_chunked_records(
        sheet,
        scene_start,
        scene_values,
        styles,
        identity_column=0,
        column_widths=[18, 24, 20, 12, 18, 24],
        minimum_height=24,
    )
    for scene_row in range(scene_start, next_row):
        sheet.cell(scene_row, 6).number_format = "0%"
    overview_continuations += scene_continuations
    row = next_row - 1

    row += 2
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    sheet.cell(row, 1, _label(language, "限制与说明", "LIMITATIONS & NOTES"))
    _section_cell(sheet.cell(row, 1), styles)
    for item in [*dataset["limitations"], *(_text_record(warning) for warning in styles["renderer_warnings"])]:
        chunks = _split_text(_plain(item))
        for chunk in chunks:
            row += 1
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            sheet.cell(row, 1, _safe_string(chunk))
            sheet.cell(row, 1).font = styles["meta_font"]
            sheet.cell(row, 1).alignment = styles["body_alignment"]
            sheet.cell(row, 1).fill = styles["warning_fill"] if not delivery["professional_export_allowed"] else styles["paper_fill"]
            sheet.row_dimensions[row].height = _row_height([chunk], minimum=24)
        overview_continuations += len(chunks) - 1

    logo_embedded = False
    logo = plan["settings"]["logo"]
    if logo is not None:
        image, stream = _bound_openpyxl_image(logo, project_root, api, "logo", max_width=118, max_height=42)
        image_streams.append(stream)
        sheet.add_image(image, "G1")
        logo_embedded = True

    widths = {"A": 18, "B": 24, "C": 20, "D": 12, "E": 18, "F": 24, "G": 15, "H": 15}
    _finish_sheet(
        sheet,
        widths,
        print_end=row,
        print_last_column="H",
        freeze=f"A{metric_start}",
        title_rows="1:1",
        project_title=_plain(project["title"]),
        styles=styles,
        api=api,
    )
    return {
        "logo_embedded": logo_embedded,
        "overview_continuation_row_count": overview_continuations,
    }


def _storyboard_sheet(
    sheet: Any,
    dataset: dict[str, Any],
    plan: dict[str, Any],
    styles: dict[str, Any],
    project_root: Path,
    api: dict[str, Any],
    image_streams: list[io.BytesIO],
) -> dict[str, Any]:
    language = plan["settings"]["language"]
    headers = [
        _label(language, "镜头号", "Shot"),
        _label(language, "叙事段落", "Section"),
        _label(language, "时间码", "Timecode"),
        _label(language, "时长（秒）", "Duration (s)"),
        _label(language, "关键帧", "Frame"),
        _label(language, "画面描述", "Visual description"),
        _label(language, "对白", "VO / Dialogue"),
        _label(language, "画面文字", "On-screen text"),
        _label(language, "音乐 / 音效", "Music / SFX"),
        _label(language, "节奏 / 转场", "Rhythm / transition"),
        _label(language, "验证状态", "Verification"),
    ]
    _table_title(sheet, _label(language, "逐镜分镜与视听证据", "SHOT-BY-SHOT AUDIOVISUAL EVIDENCE"), _status_line(dataset, plan), headers, styles)
    row = 4
    storyboard_widths = [10, 15, 18, 12, 25, 34, 28, 24, 30, 30, 28]
    continuation_rows = 0
    embedded_frames = 0
    missing_frames = 0
    for shot in dataset["shots"]:
        columns = _story_text_columns(shot, language)
        chunks = [_split_text(value) for value in columns]
        row_count = max(len(items) for items in chunks)
        start_row = row
        for part in range(row_count):
            values: list[Any] = [
                shot["shot_no"] if part == 0 else f"↳ {shot['shot_no']}.{part}",
                " / ".join(shot["scene_ids"]) if part == 0 and shot["scene_ids"] else ("—" if part == 0 else ""),
                _plain(shot["timecode"]) if part == 0 else _label(language, "续", "continued"),
                shot["duration_seconds"] if part == 0 else None,
                "",
                chunks[0][part] if part < len(chunks[0]) else "",
                chunks[1][part] if part < len(chunks[1]) else "",
                chunks[2][part] if part < len(chunks[2]) else "",
                chunks[3][part] if part < len(chunks[3]) else "",
                chunks[4][part] if part < len(chunks[4]) else "",
                chunks[5][part] if part < len(chunks[5]) else "",
            ]
            if part == 0 and not shot["frame"]["present"]:
                values[4] = _safe_string(
                    f"{_label(language, '画面证据缺失', 'Frame unavailable')}: {_plain(shot['frame']['failure'])}"
                )
            _write_row(sheet, row, values, styles, alternate=(shot["shot_no"] % 2 == 0))
            sheet.cell(row, 4).number_format = "0.00"
            sheet.row_dimensions[row].height = _row_height(
                values,
                column_widths=storyboard_widths,
                minimum=82 if part == 0 else 24,
            )
            row += 1
        continuation_rows += row_count - 1
        if shot["frame"]["present"]:
            image, stream = _bound_openpyxl_image(
                shot["frame"], project_root, api, f"frame evidence for {shot['shot_id']}", max_width=164, max_height=98
            )
            image_streams.append(stream)
            sheet.add_image(image, f"E{start_row}")
            embedded_frames += 1
        else:
            missing_frames += 1

    final_row = max(4, row - 1)
    if not dataset["shots"]:
        _write_row(sheet, 4, [_label(language, "暂无镜头证据", "No shot evidence")], styles)
    sheet.auto_filter.ref = f"A3:K{final_row}"
    widths = {"A": 10, "B": 15, "C": 18, "D": 12, "E": 25, "F": 34, "G": 28, "H": 24, "I": 30, "J": 30, "K": 28}
    _finish_sheet(
        sheet,
        widths,
        print_end=final_row,
        print_last_column="K",
        freeze="E4",
        title_rows="3:3",
        project_title=_plain(dataset["project"]["title"]),
        styles=styles,
        api=api,
    )
    return {
        "primary_shot_row_count": len(dataset["shots"]),
        "storyboard_continuation_row_count": continuation_rows,
        "embedded_frame_count": embedded_frames,
        "missing_frame_count": missing_frames,
    }


def _voice_sheet(
    sheet: Any,
    dataset: dict[str, Any],
    plan: dict[str, Any],
    styles: dict[str, Any],
    api: dict[str, Any],
) -> tuple[int, int]:
    language = plan["settings"]["language"]
    headers = [
        _label(language, "类型", "Type"),
        _label(language, "时间", "Range"),
        _label(language, "来源 ID", "Source ID"),
        _label(language, "原始提案", "Original"),
        _label(language, "有效内容", "Effective"),
        _label(language, "语言 / 角色", "Language / role"),
        _label(language, "验证 / 置信度", "Verification / confidence"),
        _label(language, "证据定位", "Evidence"),
    ]
    _table_title(sheet, _label(language, "VO、对白与画面文字", "VO, DIALOGUE & ON-SCREEN TEXT"), _status_line(dataset, plan), headers, styles)
    timed_rows: list[tuple[float, float, str, list[Any]]] = []
    for event in dataset["audio"]["events"]:
        if event["kind"] != "voice":
            continue
        original = event["original_proposal"]
        effective = event["effective_proposal"] or original
        review = event["review"]
        timed_rows.append((
            event["start_seconds"],
            event["end_seconds"],
            event["event_id"],
            [
                "VO / voice",
                _time_range(event["start_seconds"], event["end_seconds"]),
                event["event_id"],
                _plain(original["text"]),
                _plain(effective["text"]),
                f"{_plain(effective['language'])} / {effective['voice_role']}",
                " / ".join(item for item in (effective["verification"], _percent(effective["confidence"]), review["status"] if review else None) if item),
                event["evidence_reference"],
            ],
        ))
    for shot in dataset["shots"]:
        onscreen = shot["text"]["onscreen_text"]
        if onscreen["is_blank"]:
            continue
        timed_rows.append((
            shot["start_seconds"],
            shot["end_seconds"],
            shot["shot_id"],
            [
                _label(language, "画面文字", "On-screen text"),
                _plain(shot["timecode"]),
                shot["shot_id"],
                _plain(onscreen),
                _plain(onscreen),
                dataset["project"]["delivery_language"],
                shot["verification"]["annotation_verification"],
                shot["evidence_reference"],
            ],
        ))
    timed_rows.sort(key=lambda item: (item[0], item[1], item[2]))
    rows = [item[3] for item in timed_rows]
    if not rows:
        rows = [[_label(language, "暂无 VO / 画面文字证据", "No VO or on-screen text evidence")]]
    primary_count = len(rows) if rows and len(rows[0]) > 1 else 0
    voice_widths = [18, 20, 20, 38, 38, 22, 26, 38]
    next_row, continuations = _write_chunked_records(
        sheet,
        4,
        rows,
        styles,
        identity_column=2,
        column_widths=voice_widths,
        minimum_height=30,
    )
    final_row = next_row - 1
    sheet.auto_filter.ref = f"A3:H{final_row}"
    widths = {"A": 18, "B": 20, "C": 20, "D": 38, "E": 38, "F": 22, "G": 26, "H": 38}
    _finish_sheet(sheet, widths, print_end=final_row, print_last_column="H", freeze="D4", title_rows="3:3", project_title=_plain(dataset["project"]["title"]), styles=styles, api=api)
    return primary_count, continuations


def _audio_sheet(
    sheet: Any,
    dataset: dict[str, Any],
    plan: dict[str, Any],
    styles: dict[str, Any],
    api: dict[str, Any],
) -> tuple[int, bool, int, int, str, int]:
    language = plan["settings"]["language"]
    headers = [
        "Event ID",
        _label(language, "类型", "Kind"),
        _label(language, "开始（秒）", "Start (s)"),
        _label(language, "结束（秒）", "End (s)"),
        _label(language, "时长（秒）", "Duration (s)"),
        _label(language, "标签 / 内容", "Label / text"),
        _label(language, "能量", "Energy"),
        _label(language, "Onset 密度", "Onset density"),
        _label(language, "BPM（估计）", "Estimated BPM"),
        _label(language, "身份 / 审核", "Identity / review"),
        _label(language, "关联镜头", "Linked shots"),
        _label(language, "证据定位", "Evidence"),
    ]
    _table_title(sheet, _label(language, "音乐、音效与节奏时间线", "MUSIC, SFX & RHYTHM TIMELINE"), _status_line(dataset, plan), headers, styles)
    shot_links: dict[str, list[str]] = {event["event_id"]: [] for event in dataset["audio"]["events"]}
    for shot in dataset["shots"]:
        for link in shot["audio"]["event_links"]:
            shot_links[link["event_id"]].append(shot["shot_id"])
    rows: list[list[Any]] = []
    for event in dataset["audio"]["events"]:
        effective = event["effective_proposal"] or event["original_proposal"]
        label_text = " — ".join(item for item in (_plain(effective["label"]), _plain(effective["text"])) if item)
        review = event["review"]
        values = [
            event["event_id"],
            event["kind"],
            event["start_seconds"],
            event["end_seconds"],
            event["end_seconds"] - event["start_seconds"],
            label_text,
            effective["energy"],
            effective["onset_density"],
            effective["estimated_bpm"],
            " / ".join(item for item in (event["identity_status"], effective["verification"], review["status"] if review else None) if item),
            ", ".join(shot_links[event["event_id"]]) or "—",
            event["evidence_reference"],
        ]
        rows.append(values)
    if not dataset["audio"]["events"]:
        rows = [[_label(language, "音频时间线不可用；这不是静音证据", "Audio timeline unavailable; this is not evidence of silence")]]
    audio_widths = [22, 12, 12, 12, 12, 42, 12, 14, 14, 25, 24, 42]
    next_row, continuations = _write_chunked_records(
        sheet,
        4,
        rows,
        styles,
        identity_column=0,
        column_widths=audio_widths,
        minimum_height=24,
    )
    final_row = next_row - 1
    for row in range(4, final_row + 1):
        for column in (3, 4, 5, 7, 8, 9):
            sheet.cell(row, column).number_format = "0.00"
    sheet.auto_filter.ref = f"A3:L{final_row}"
    chart_added, chart_source_points, chart_points, chart_method = _add_energy_chart(
        sheet,
        dataset["audio"]["events"],
        final_row,
        language,
        api,
        styles,
    )
    print_end = final_row + 18 if chart_added else final_row + (2 if chart_method == "omitted_over_limit" else 0)
    widths = {"A": 22, "B": 12, "C": 12, "D": 12, "E": 12, "F": 42, "G": 12, "H": 14, "I": 14, "J": 25, "K": 24, "L": 42}
    _finish_sheet(sheet, widths, print_end=print_end, print_last_column="L", freeze="F4", title_rows="3:3", project_title=_plain(dataset["project"]["title"]), styles=styles, api=api)
    return (
        len(dataset["audio"]["events"]),
        chart_added,
        chart_source_points,
        chart_points,
        chart_method,
        continuations,
    )


def _evidence_sheet(
    sheet: Any,
    dataset: dict[str, Any],
    plan: dict[str, Any],
    styles: dict[str, Any],
    api: dict[str, Any],
) -> tuple[int, int]:
    language = plan["settings"]["language"]
    headers = [
        _label(language, "类别", "Category"),
        _label(language, "项目", "Item"),
        _label(language, "状态", "Status"),
        _label(language, "说明", "Details"),
        _label(language, "证据定位", "Evidence"),
    ]
    _table_title(sheet, _label(language, "证据、能力与限制说明", "EVIDENCE, CAPABILITIES & LIMITATIONS"), _status_line(dataset, plan), headers, styles)
    rows: list[list[Any]] = [
        ["binding", "dataset", "bound", dataset["dataset_digest"], "data/client_export_dataset.json"],
        ["binding", "template", "bound", f"{plan['template']['template_id']} {plan['template']['template_version']} / {plan['template']['template_digest']}", "templates/client/manifest.json"],
        ["delivery", "readiness", dataset["delivery_status"]["readiness_status"], dataset["delivery_status"]["state"], dataset["delivery_status"]["readiness_reference"]],
    ]
    for key, value in dataset["field_semantics"].items():
        rows.append(["semantics", key, "defined", value, "client-export-dataset/v1"])
    for name, capability in dataset["audio"]["capabilities"].items():
        rows.append(["capability", name, capability["status"], _plain(capability["reason"]), capability["source_id"] or "—"])
    for source in dataset["audio"]["sources"]:
        details = " / ".join(item for item in (_plain(source["adapter"]), _plain(source["engine"]), _plain(source["model"])) if item)
        rows.append(["source", source["capability"], source["status"], details, source["source_id"]])
    for item in dataset["limitations"]:
        rows.append(["limitation", "client-safe", "visible", _plain(item), "—"])
    for item in dataset["unresolved_items"]:
        rows.append(["unresolved", item["scope"], "open", _plain(item["reason"]), item["evidence_reference"] or "—"])
    evidence_widths = [18, 24, 18, 70, 46]
    next_row, continuations = _write_chunked_records(
        sheet,
        4,
        rows,
        styles,
        identity_column=1,
        column_widths=evidence_widths,
        minimum_height=24,
    )
    final_row = next_row - 1
    sheet.auto_filter.ref = f"A3:E{final_row}"
    widths = {"A": 18, "B": 24, "C": 18, "D": 70, "E": 46}
    _finish_sheet(sheet, widths, print_end=final_row, print_last_column="E", freeze="D4", title_rows="3:3", project_title=_plain(dataset["project"]["title"]), styles=styles, api=api)
    return len(rows), continuations


def _table_title(sheet: Any, title: str, subtitle: str, headers: list[str], styles: dict[str, Any]) -> None:
    last = len(headers)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last)
    sheet.cell(1, 1, _safe_string(title))
    sheet.cell(1, 1).font = styles["title_font"]
    sheet.cell(1, 1).alignment = styles["title_alignment"]
    for cell in sheet[1][:last]:
        cell.fill = styles["title_fill"]
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last)
    sheet.cell(2, 1, _safe_string(subtitle))
    sheet.cell(2, 1).font = styles["subtitle_font"]
    sheet.cell(2, 1).alignment = styles["title_alignment"]
    for cell in sheet[2][:last]:
        cell.fill = styles["title_fill"]
    for column, value in enumerate(headers, 1):
        sheet.cell(3, column, _safe_string(value))
        _header_cell(sheet.cell(3, column), styles)
    sheet.row_dimensions[1].height = 34
    sheet.row_dimensions[2].height = 22
    sheet.row_dimensions[3].height = 32


def _finish_sheet(
    sheet: Any,
    widths: dict[str, float],
    *,
    print_end: int,
    print_last_column: str,
    freeze: str,
    title_rows: str,
    project_title: str,
    styles: dict[str, Any],
    api: dict[str, Any] | None,
) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = freeze
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 85
    sheet.print_title_rows = title_rows
    sheet.print_area = f"A1:{print_last_column}{max(1, print_end)}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    if api is not None:
        sheet.page_margins = api["PageMargins"](left=0.35, right=0.35, top=0.55, bottom=0.55, header=0.2, footer=0.25)
    sheet.print_options.horizontalCentered = True
    sheet.oddFooter.left.text = _header_footer_text(project_title[:40])
    sheet.oddFooter.right.text = "Page &P / &N"
    sheet.oddFooter.left.size = 8
    sheet.oddFooter.right.size = 8


def _story_text_columns(shot: dict[str, Any], language: str) -> list[str]:
    text = shot["text"]
    camera = shot["camera"]
    visual = _join_unique([
        _localized_pair(text["content_summary_zh"], text["content_summary"], language),
        _plain(text["visual_description"]),
        _localized_pair(text["subject_zh"], text["subject"], language),
        _localized_pair(text["action_zh"], text["action"], language),
        _join_unique([_plain(camera["shot_scale"]), _plain(camera["angle"]), _plain(camera["motion"]), _plain(camera["composition"])], separator=" · "),
    ])
    voice = _join_unique([_plain(text["dialogue"]), _plain(text["speech_summary"])])
    onscreen = _plain(text["onscreen_text"])
    music = _join_unique([_plain(text["music_state"]), _plain(text["sound_design"])])
    rhythm = _join_unique([
        _plain(text["rhythm_notes"]),
        _plain(text["sound_rhythm"]),
        f"{_plain(text['transition_in'])} → {_plain(text['transition_out'])}",
    ])
    verification = _join_unique([
        f"{shot['verification']['readiness_status']} / {shot['verification']['annotation_verification']}",
        f"confidence {_percent(shot['verification']['visual_confidence'])}",
        *(_plain(item) for item in shot["verification"]["readiness_reasons"]),
        shot["evidence_reference"],
    ])
    return [visual, voice, onscreen, music, rhythm, verification]


def _localized_pair(zh: dict[str, Any], en: dict[str, Any], language: str) -> str:
    if language == "zh":
        return _plain(zh) or _plain(en)
    if language == "en":
        return _plain(en) or _plain(zh)
    return _join_unique([_plain(zh), _plain(en)])


def _bound_openpyxl_image(
    receipt: dict[str, Any],
    project_root: Path,
    api: dict[str, Any],
    label: str,
    *,
    max_width: int,
    max_height: int,
) -> tuple[Any, io.BytesIO]:
    path = receipt.get("path")
    if type(path) is not str:
        raise XlsxExportError(f"{label} path is invalid")
    candidate = (project_root / path).resolve()
    try:
        candidate.relative_to(project_root)
        raw = read_regular_bytes(candidate, root=project_root, max_bytes=16 * 1024 * 1024)
        actual = inspect_image_bytes(raw, max_bytes=16 * 1024 * 1024).receipt_fields()
    except (OSError, ValueError) as exc:
        raise XlsxExportError(f"{label} is unavailable or unsafe") from exc
    expected = {key: receipt.get(key) for key in ("sha256", "size_bytes", "media_type", "width", "height")}
    if actual != expected:
        raise XlsxExportError(f"{label} changed after dataset binding")
    stream = io.BytesIO(raw)
    image = api["Image"](stream)
    scale = min(1.0, max_width / max(1, image.width), max_height / max(1, image.height))
    image.width = max(1, round(image.width * scale))
    image.height = max(1, round(image.height * scale))
    return image, stream


def _add_energy_chart(
    sheet: Any,
    events: list[dict[str, Any]],
    final_row: int,
    language: str,
    api: dict[str, Any],
    styles: dict[str, Any],
) -> tuple[bool, int, int, str]:
    points = [
        (event["start_seconds"], (event["effective_proposal"] or event["original_proposal"])["energy"])
        for event in events
        if (event["effective_proposal"] or event["original_proposal"])["energy"] is not None
    ]
    if not points:
        return False, 0, 0, "none"
    if len(points) > 240:
        note_row = final_row + 2
        sheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=12)
        sheet.cell(
            note_row,
            1,
            _label(
                language,
                f"音频能量图已省略：{len(points)} 个测量点超过 240 点可视化上限；完整事件仍保留在表格中。",
                f"Audio energy chart omitted: {len(points)} measured points exceed the 240-point visual limit; the full event table is preserved.",
            ),
        )
        sheet.cell(note_row, 1).font = styles["warning_font"]
        sheet.cell(note_row, 1).fill = styles["warning_fill"]
        sheet.cell(note_row, 1).alignment = styles["body_alignment"]
        sheet.row_dimensions[note_row].height = 30
        return False, len(points), 0, "omitted_over_limit"
    sheet.cell(3, 14, "Chart time (s)")
    sheet.cell(3, 15, "Chart energy")
    _header_cell(sheet.cell(3, 14), styles)
    _header_cell(sheet.cell(3, 15), styles)
    for row, (start, energy) in enumerate(points, 4):
        sheet.cell(row, 14, start)
        sheet.cell(row, 15, energy)
        for column in (14, 15):
            sheet.cell(row, column).font = styles["meta_font"]
            sheet.cell(row, column).number_format = "0.00"
    sheet.column_dimensions["N"].width = 14
    sheet.column_dimensions["O"].width = 14
    last = 3 + len(points)
    chart = api["ScatterChart"]()
    chart.title = _label(language, "音频能量", "Audio energy")
    chart.style = 13
    chart.height = 5.5
    chart.width = 12
    chart.y_axis.title = "Energy"
    chart.x_axis.title = _label(language, "开始时间（秒）", "Start time (s)")
    y_values = api["Reference"](sheet, min_col=15, min_row=4, max_row=last)
    x_values = api["Reference"](sheet, min_col=14, min_row=4, max_row=last)
    chart.series.append(api["Series"](y_values, x_values, title="Energy"))
    chart.scatterStyle = "line"
    chart.visible_cells_only = False
    chart.legend = None
    sheet.column_dimensions["N"].hidden = True
    sheet.column_dimensions["O"].hidden = True
    sheet.add_chart(chart, f"A{final_row + 2}")
    return True, len(points), len(points), "scatter_all"


def _validate_xlsx_archive(path: Path, *, expected_images: int) -> dict[str, Any]:
    try:
        with zipfile.ZipFile(path) as archive:
            if archive.testzip() is not None:
                raise XlsxExportError("XLSX ZIP integrity check failed")
            names = set(archive.namelist())
            required = {"[Content_Types].xml", "xl/workbook.xml", "xl/styles.xml"}
            if not required.issubset(names):
                raise XlsxExportError("XLSX package is incomplete")
            # Every XML member below was emitted moments earlier by this
            # bounded renderer; the product has no arbitrary-workbook ingest.
            workbook_xml = ElementTree.fromstring(  # nosec B314
                archive.read("xl/workbook.xml")
            )
            sheet_names = [item.attrib.get("name") for item in workbook_xml.findall(".//{*}sheet")]
            if sheet_names != list(SHEET_NAMES):
                raise XlsxExportError("XLSX sheet order does not match the template")
            formula_count = 0
            for name in names:
                if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                    root = ElementTree.fromstring(archive.read(name))  # nosec B314
                    formula_count += len(root.findall(".//{*}f"))
            external_links = sum(name.startswith("xl/externalLinks/") for name in names)
            for name in names:
                if not name.endswith(".rels"):
                    continue
                relationships = ElementTree.fromstring(  # nosec B314
                    archive.read(name)
                )
                external_links += sum(item.attrib.get("TargetMode") == "External" for item in relationships.findall(".//{*}Relationship"))
            contains_macros = "xl/vbaProject.bin" in names or any("macroEnabled" in archive.read(name).decode("utf-8", "ignore") for name in names if name == "[Content_Types].xml")
            images = sum(name.startswith("xl/media/") and not name.endswith("/") for name in names)
    except (OSError, zipfile.BadZipFile, ElementTree.ParseError) as exc:
        raise XlsxExportError("XLSX package validation failed") from exc
    if formula_count or external_links or contains_macros:
        raise XlsxExportError("XLSX package contains active or external content")
    if images != expected_images:
        raise XlsxExportError("XLSX embedded image count does not match the renderer receipt")
    return {
        "sheet_count": len(sheet_names),
        "formula_count": formula_count,
        "external_link_count": external_links,
        "contains_macros": contains_macros,
    }


def _write_chunked_records(
    sheet: Any,
    start_row: int,
    records: list[list[Any]],
    styles: dict[str, Any],
    *,
    identity_column: int,
    column_widths: list[float],
    minimum_height: float,
) -> tuple[int, int]:
    row = start_row
    continuation_count = 0
    for record_index, values in enumerate(records):
        chunks = [
            _split_text(value) if isinstance(value, str) else [value]
            for value in values
        ]
        record_rows = max(len(items) for items in chunks)
        for part in range(record_rows):
            rendered = [items[part] if part < len(items) else "" for items in chunks]
            if part and identity_column < len(rendered):
                identity = str(values[identity_column] or record_index + 1)
                rendered[identity_column] = f"↳ {identity}"
            _write_row(sheet, row, rendered, styles, alternate=(record_index % 2 == 1))
            sheet.row_dimensions[row].height = _row_height(
                rendered,
                column_widths=column_widths,
                minimum=minimum_height,
            )
            row += 1
        continuation_count += record_rows - 1
    return row, continuation_count


def _write_row(sheet: Any, row: int, values: list[Any], styles: dict[str, Any], *, alternate: bool = False) -> None:
    if row > EXCEL_MAX_ROWS:
        raise XlsxExportError("XLSX sheet exceeds Excel's row limit")
    for column, value in enumerate(values, 1):
        cell = sheet.cell(row, column, _safe_scalar(value))
        cell.font = styles["body_font"]
        cell.alignment = styles["body_alignment"]
        cell.border = styles["thin_bottom"]
        if alternate:
            cell.fill = styles["paper_fill"]


def _header_cell(cell: Any, styles: dict[str, Any]) -> None:
    cell.font = styles["header_font"]
    cell.fill = styles["header_fill"]
    cell.alignment = styles["center_alignment"]
    cell.border = styles["section_border"]


def _section_cell(cell: Any, styles: dict[str, Any]) -> None:
    cell.font = styles["header_font"]
    cell.fill = styles["accent_fill"]
    cell.alignment = styles["title_alignment"]
    cell.border = styles["section_border"]


def _split_text(value: str) -> list[str]:
    text = _unbounded_string(value)
    if not text:
        return [""]
    chunks: list[str] = []
    current: list[str] = []
    line_units = 0.0
    lines = 1
    previous_cr = False
    for character in text:
        if character == "\n" and previous_cr:
            previous_cr = False
            continue
        if character in "\r\n":
            if lines >= XLSX_MAX_LINES_PER_ROW:
                chunks.append("".join(current))
                current = []
                lines = 1
            else:
                current.append("\n")
                lines += 1
            line_units = 0.0
            previous_cr = character == "\r"
            continue
        previous_cr = False
        units = 1.75 if CJK_RE.match(character) else 1.0
        if line_units + units > 42:
            if lines >= XLSX_MAX_LINES_PER_ROW:
                chunks.append("".join(current))
                current = []
                lines = 1
            else:
                lines += 1
            line_units = 0.0
        current.append(character)
        line_units += units
        if len(current) >= 32_000:
            chunks.append("".join(current))
            current = []
            lines = 1
            line_units = 0.0
    if current or not chunks:
        chunks.append("".join(current))
    if any(len(chunk) > EXCEL_CELL_LIMIT for chunk in chunks):
        raise XlsxExportError("renderer text exceeds Excel's cell limit")
    return chunks


def _row_height(
    values: list[Any],
    *,
    column_widths: list[float] | None = None,
    minimum: float = 22,
) -> float:
    lines = 1
    for index, value in enumerate(values):
        if value is None:
            continue
        text = str(value)
        width = column_widths[index] if column_widths is not None and index < len(column_widths) else 42
        capacity = max(6.0, width - 2)
        estimated = 0
        for logical_line in re.split(r"\r\n|\r|\n", text):
            units = sum(1.75 if CJK_RE.match(character) else 1.0 for character in logical_line)
            estimated += max(1, math.ceil(units / capacity))
        lines = max(lines, estimated)
    lines = min(XLSX_MAX_LINES_PER_ROW, lines)
    return min(320, max(minimum, lines * 15))


def _plain(cell: dict[str, Any]) -> str:
    value = cell.get("spreadsheet_text")
    if type(value) is not str:
        raise XlsxExportError("renderer received an invalid text cell")
    return _unbounded_string(value)


def _text_record(value: str) -> dict[str, str]:
    return {"spreadsheet_text": value}


def _safe_scalar(value: Any) -> str | int | float | bool | None:
    if value is None or type(value) in {int, float, bool}:
        return value
    return _safe_string(str(value))


def _safe_string(value: str) -> str:
    value = _unbounded_string(value)
    if len(value) > EXCEL_CELL_LIMIT:
        raise XlsxExportError("renderer text exceeds Excel's cell limit")
    return value


def _unbounded_string(value: str) -> str:
    if type(value) is not str or ILLEGAL_XML_RE.search(value):
        raise XlsxExportError("renderer text contains unsupported XML control characters")
    if value.startswith(("=", "+", "-", "@", "\t")):
        return "'" + value
    return value


def _header_footer_text(value: str) -> str:
    return _safe_string(value).replace("&", "&&")


def _join_unique(values: list[str], *, separator: str = "\n") -> str:
    result: list[str] = []
    for value in values:
        if value and value not in result:
            result.append(value)
    return separator.join(result)


def _time_range(start: float, end: float) -> str:
    return f"{start:.3f}–{end:.3f}"


def _percent(value: float | None) -> str:
    return "unknown" if value is None else f"{value:.0%}"


def _label(language: str, zh: str, en: str) -> str:
    if language == "zh":
        return zh
    if language == "en":
        return en
    return f"{zh} / {en}"


def _status_line(dataset: dict[str, Any], plan: dict[str, Any]) -> str:
    delivery = dataset["delivery_status"]
    language = plan["settings"]["language"]
    status = _label(language, "专业可交付" if delivery["professional_export_allowed"] else "仅供草稿审核", "PROFESSIONAL" if delivery["professional_export_allowed"] else "DRAFT ONLY")
    return _safe_string(
        f"{_plain(dataset['project']['title'])} · {status} · dataset {dataset['dataset_digest'][:12]} · template {plan['template']['template_version']}"
    )


__all__ = ["XlsxExportError", "render_client_xlsx"]
