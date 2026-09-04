from __future__ import annotations

import copy
import importlib.util
import tempfile
import unittest
import zipfile
from datetime import datetime
from pathlib import Path
from unittest.mock import patch

from tests.test_audio_review import audio_review_fixture
from tests.test_evidence_handoff import PNG_1X1
from video_analysis_mvp.client_export_dataset import (
    _canonical_digest,
    _client_text,
    build_client_export_dataset,
)
from video_analysis_mvp.export_xlsx import XlsxExportError, render_client_xlsx
from video_analysis_mvp.image_evidence import inspect_image_bytes

OPENPYXL_AVAILABLE = importlib.util.find_spec("openpyxl") is not None
if OPENPYXL_AVAILABLE:
    from openpyxl import load_workbook


@unittest.skipUnless(OPENPYXL_AVAILABLE, "install the export extra to run XLSX renderer tests")
class XlsxExportTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory(prefix="vew-xlsx-")
        self.addCleanup(self.temp.cleanup)
        self.paths = audio_review_fixture(Path(self.temp.name))
        self.dataset = build_client_export_dataset(self.paths)

    def redigest(self) -> None:
        base = {key: value for key, value in self.dataset.items() if key not in {"dataset_id", "dataset_digest"}}
        digest = _canonical_digest(base)
        self.dataset["dataset_id"] = self.dataset["dataset_digest"] = digest

    def output(self, name: str = "client-breakdown.xlsx") -> Path:
        return self.paths.root / name

    def workbook_text(self, workbook) -> list[str]:
        return [
            value
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if isinstance((value := cell.value), str)
        ]

    def test_fixed_five_sheet_contract_is_safe_and_client_readable(self) -> None:
        unsafe = '=HYPERLINK("https://example.invalid","click")'
        self.dataset["shots"][0]["text"]["content_summary"] = _client_text(unsafe, "test summary")
        self.redigest()
        target = self.output()

        receipt = render_client_xlsx(
            self.dataset,
            target,
            settings={"language": "bilingual", "project_subtitle": "Campaign evidence review"},
            project_root=self.paths.root,
            available_fonts=["PingFang SC"],
        )

        self.assertEqual("xlsx-render-receipt/v1", receipt["schema_id"])
        self.assertEqual(self.dataset["dataset_digest"], receipt["dataset_digest"])
        self.assertEqual("client-storyboard", receipt["template_id"])
        self.assertEqual(5, receipt["sheet_count"])
        self.assertEqual(0, receipt["formula_count"])
        self.assertEqual(0, receipt["external_link_count"])
        self.assertFalse(receipt["contains_macros"])
        self.assertEqual("PingFang SC", receipt["font_plan"]["selected_cjk_font"])
        self.assertEqual("Noto Sans CJK SC", receipt["declared_font_name"])
        self.assertFalse(receipt["declared_font_verified"])
        self.assertTrue(any("font" in warning.lower() for warning in receipt["warnings"]))
        self.assertRegex(receipt["sha256"], r"^[0-9a-f]{64}$")
        self.assertEqual(receipt["size_bytes"], target.stat().st_size)
        generated_at = datetime.fromisoformat(receipt["generated_at_utc"])
        self.assertIsNotNone(generated_at.tzinfo)

        workbook = load_workbook(target, data_only=False, keep_links=False)
        self.addCleanup(workbook.close)
        self.assertEqual(
            ["01_项目概览", "02_逐镜分镜表", "03_VO与画面文字", "04_音乐与节奏", "05_证据与说明"],
            workbook.sheetnames,
        )
        self.assertTrue(all(sheet.sheet_state == "visible" for sheet in workbook.worksheets))
        self.assertEqual(generated_at.replace(tzinfo=None), workbook.properties.created)
        self.assertTrue(all(sheet.page_margins.left == 0.35 for sheet in workbook.worksheets))
        self.assertTrue(all(cell.data_type != "f" for sheet in workbook.worksheets for row in sheet.iter_rows() for cell in row))
        story = workbook["02_逐镜分镜表"]
        overview = workbook["01_项目概览"]
        self.assertEqual("E4", story.freeze_panes)
        self.assertEqual("A3:K4", story.auto_filter.ref)
        self.assertIn("$3:$3", str(story.print_title_rows))
        self.assertIsNone(story.print_title_cols)
        self.assertIn("$A$1:$K$4", str(story.print_area))
        self.assertEqual("landscape", story.page_setup.orientation)
        self.assertEqual(1, story.page_setup.fitToWidth)
        self.assertEqual(0, overview.page_setup.fitToHeight)
        self.assertIn("$1:$1", str(overview.print_title_rows))
        self.assertGreaterEqual(overview.column_dimensions["D"].width, 12)
        self.assertEqual("Noto Sans CJK SC", story["A1"].font.name)
        self.assertEqual("2.00 s", overview["B5"].value)
        self.assertEqual("1.778:1", overview["B7"].value)
        self.assertIn("镜头号 / Shot", story["A3"].value)
        self.assertTrue(any(value.startswith("'=") for value in self.workbook_text(workbook)))
        self.assertFalse(any(str(self.paths.root) in value for value in self.workbook_text(workbook)))
        self.assertFalse(any("{'en':" in value for value in self.workbook_text(workbook)))

        with zipfile.ZipFile(target) as archive:
            names = set(archive.namelist())
            self.assertNotIn("xl/vbaProject.bin", names)
            self.assertFalse(any(name.startswith("xl/externalLinks/") for name in names))

    def test_present_frame_is_digest_bound_and_missing_frame_stays_explicit(self) -> None:
        frame_path = self.paths.keyframes / "client-frame.png"
        frame_path.write_bytes(PNG_1X1)
        evidence = inspect_image_bytes(PNG_1X1).receipt_fields()
        first = self.dataset["shots"][0]
        first["frame"] = {
            "path": "assets/keyframes/client-frame.png",
            "present": True,
            **evidence,
            "failure": _client_text("", "frame failure"),
        }
        second = copy.deepcopy(first)
        second.update(
            shot_id="shot_0002",
            shot_no=2,
            start_seconds=2.0,
            end_seconds=3.0,
            duration_seconds=1.0,
            timecode=_client_text("00:02-00:03", "timecode"),
            evidence_reference="data/shots.json#shot_id=shot_0002",
        )
        second["frame"] = {
            "path": "assets/keyframes/missing.png",
            "present": False,
            "sha256": None,
            "size_bytes": None,
            "media_type": None,
            "width": None,
            "height": None,
            "failure": _client_text("Primary frame unavailable", "frame failure"),
        }
        second["audio"]["event_links"] = []
        second["audio"]["event_coverage_seconds"] = {kind: 0.0 for kind in ("voice", "music", "sfx", "silence", "mixed")}
        self.dataset["shots"] = [first, second]
        self.dataset["project"]["duration_seconds"] = 3.0
        self.redigest()

        target = self.output()
        receipt = render_client_xlsx(self.dataset, target, project_root=self.paths.root)
        self.assertEqual(1, receipt["embedded_image_count"])
        self.assertEqual(1, receipt["missing_frame_count"])
        workbook = load_workbook(target, keep_links=False)
        self.addCleanup(workbook.close)
        story = workbook["02_逐镜分镜表"]
        self.assertEqual(1, len(story._images))
        self.assertEqual(1, story._images[0].width)
        self.assertEqual(1, story._images[0].height)
        self.assertIn("Primary frame unavailable", story["E5"].value)

        frame_path.write_bytes(PNG_1X1 + b"tampered")
        with self.assertRaisesRegex(XlsxExportError, "frame evidence"):
            render_client_xlsx(self.dataset, self.output("tampered.xlsx"), project_root=self.paths.root)
        self.assertFalse(self.output("tampered.xlsx").exists())

    def test_long_text_uses_bounded_continuation_rows_without_excel_cell_overflow(self) -> None:
        long_text = "a" * 60_000
        self.dataset["shots"][0]["text"]["content_summary"] = _client_text(long_text, "long summary")
        self.redigest()

        receipt = render_client_xlsx(self.dataset, self.output(), project_root=self.paths.root)
        self.assertGreater(receipt["continuation_row_count"], 0)
        workbook = load_workbook(self.output(), keep_links=False)
        self.addCleanup(workbook.close)
        story = workbook["02_逐镜分镜表"]
        primary_rows = [row for row in range(4, story.max_row + 1) if isinstance(story.cell(row, 1).value, int)]
        continuation_rows = [row for row in range(4, story.max_row + 1) if str(story.cell(row, 1).value or "").startswith("↳")]
        self.assertEqual([4], primary_rows)
        self.assertGreater(len(continuation_rows), 1)
        self.assertTrue(
            all(len(value) <= 32_000 for value in self.workbook_text(workbook)),
            "no generated string may exceed Excel's cell limit",
        )
        self.assertTrue(all((story.row_dimensions[row].height or 0) <= 320 for row in continuation_rows))

    def test_two_hundred_five_shots_keep_primary_rows_filters_and_print_settings(self) -> None:
        source = self.dataset["shots"][0]
        shots = []
        zero_coverage = {kind: 0.0 for kind in ("voice", "music", "sfx", "silence", "mixed")}
        for index in range(1, 206):
            shot = copy.deepcopy(source)
            shot.update(
                shot_id=f"shot_{index:04d}",
                shot_no=index,
                start_seconds=float(index - 1),
                end_seconds=float(index),
                duration_seconds=1.0,
                timecode=_client_text(f"{index - 1:04d}-{index:04d}", "timecode"),
                evidence_reference=f"data/shots.json#shot_id=shot_{index:04d}",
            )
            shot["frame"] = {
                "path": f"assets/keyframes/frame-{index:04d}.jpg",
                "present": False,
                "sha256": None,
                "size_bytes": None,
                "media_type": None,
                "width": None,
                "height": None,
                "failure": _client_text("Primary frame unavailable", "frame failure"),
            }
            shot["audio"]["event_links"] = []
            shot["audio"]["event_coverage_seconds"] = dict(zero_coverage)
            shots.append(shot)
        self.dataset["shots"] = shots
        self.dataset["project"]["duration_seconds"] = 205.0
        self.redigest()

        receipt = render_client_xlsx(self.dataset, self.output(), project_root=self.paths.root)
        self.assertEqual(205, receipt["primary_shot_row_count"])
        self.assertEqual(205, receipt["missing_frame_count"])
        workbook = load_workbook(self.output(), read_only=False, keep_links=False)
        self.addCleanup(workbook.close)
        story = workbook["02_逐镜分镜表"]
        self.assertEqual(205, sum(isinstance(story.cell(row, 1).value, int) for row in range(4, story.max_row + 1)))
        self.assertEqual(f"A3:K{story.max_row}", story.auto_filter.ref)
        self.assertIn(f"$A$1:$K${story.max_row}", str(story.print_area))
        self.assertLess(self.output().stat().st_size, 5 * 1024 * 1024)

    def test_long_subtitle_audio_and_limitation_text_continue_without_loss(self) -> None:
        limitation = "b" * 40_000
        voice = "d" * 40_000
        scene_function = "e" * 60_000
        self.dataset["limitations"][0] = _client_text(limitation, "long limitation")
        event = next(item for item in self.dataset["audio"]["events"] if item["kind"] == "voice")
        event["original_proposal"]["text"] = _client_text(voice, "long voice")
        event["effective_proposal"]["text"] = _client_text(voice, "long voice")
        self.dataset["scenes"] = [{
            "scene_id": "scene-0001",
            "start_seconds": 0.0,
            "end_seconds": 2.0,
            "shot_ids": ["shot_0001"],
            "function": _client_text(scene_function, "long scene function"),
            "pace": _client_text("measured", "scene pace"),
            "confidence": 0.8,
        }]
        self.dataset["shots"][0]["scene_ids"] = ["scene-0001"]
        self.redigest()
        subtitle = "c" * 40_000

        receipt = render_client_xlsx(
            self.dataset,
            self.output(),
            settings={"project_subtitle": subtitle},
            project_root=self.paths.root,
        )

        self.assertGreater(receipt["auxiliary_continuation_row_count"], 0)
        workbook = load_workbook(self.output(), keep_links=False)
        self.addCleanup(workbook.close)
        values = self.workbook_text(workbook)
        self.assertGreaterEqual(sum(value.count("b") for value in values), len(limitation))
        self.assertGreaterEqual(sum(value.count("c") for value in values), len(subtitle))
        self.assertGreaterEqual(sum(value.count("d") for value in values), len(voice))
        self.assertGreaterEqual(sum(value.count("e") for value in values), len(scene_function))
        self.assertTrue(all(len(value) <= 32_000 for value in values))
        self.assertEqual(0, workbook["01_项目概览"].page_setup.fitToHeight)

    def test_audio_and_evidence_sheets_preserve_counts_and_review_boundaries(self) -> None:
        receipt = render_client_xlsx(self.dataset, self.output(), project_root=self.paths.root)
        self.assertEqual(len(self.dataset["audio"]["events"]), receipt["audio_event_row_count"])
        workbook = load_workbook(self.output(), keep_links=False)
        self.addCleanup(workbook.close)
        voice = workbook["03_VO与画面文字"]
        rhythm = workbook["04_音乐与节奏"]
        evidence = workbook["05_证据与说明"]
        voice_text = "\n".join(value for row in voice.iter_rows(values_only=True) for value in row if isinstance(value, str))
        rhythm_text = "\n".join(value for row in rhythm.iter_rows(values_only=True) for value in row if isinstance(value, str))
        evidence_text = "\n".join(value for row in evidence.iter_rows(values_only=True) for value in row if isinstance(value, str))
        self.assertIn("Original VO", voice_text)
        self.assertIn("machine_estimated", voice_text)
        self.assertIn("silence-000000", rhythm_text)
        self.assertIn("final mix", evidence_text)
        self.assertIn("classification", evidence_text)
        self.assertIn(self.dataset["dataset_digest"], evidence_text)
        self.assertEqual("A3:H4", voice.auto_filter.ref)
        self.assertEqual(f"A3:L{rhythm.max_row}", rhythm.auto_filter.ref)
        self.assertGreaterEqual(voice.row_dimensions[4].height, 24)
        self.assertIn("$3:$3", str(evidence.print_title_rows))
        self.assertGreaterEqual(evidence.row_dimensions[5].height, 30)

    def test_audio_chart_uses_bounded_support_rows_after_event_continuations(self) -> None:
        event = self.dataset["audio"]["events"][0]
        long_label = "energy " + ("x" * 60_000)
        event["original_proposal"]["label"] = _client_text(long_label, "long energy label")
        event["effective_proposal"]["label"] = _client_text(long_label, "long energy label")
        self.redigest()

        receipt = render_client_xlsx(self.dataset, self.output(), project_root=self.paths.root)
        self.assertGreater(receipt["auxiliary_continuation_row_count"], 0)
        self.assertGreater(receipt["audio_visualization_point_count"], 0)
        self.assertLessEqual(receipt["audio_visualization_point_count"], 240)
        self.assertEqual(4, receipt["audio_visualization_source_point_count"])
        self.assertEqual("scatter_all", receipt["audio_visualization_method"])
        workbook = load_workbook(self.output(), keep_links=False)
        self.addCleanup(workbook.close)
        rhythm = workbook["04_音乐与节奏"]
        self.assertEqual("Chart time (s)", rhythm["N3"].value)
        self.assertEqual("Chart energy", rhythm["O3"].value)
        value_formula = rhythm._charts[0].series[0].yVal.numRef.f
        category_formula = rhythm._charts[0].series[0].xVal.numRef.f
        self.assertIn("$O$4:$O$", value_formula)
        self.assertIn("$N$4:$N$", category_formula)
        self.assertFalse(rhythm._charts[0].visible_cells_only)
        self.assertTrue(rhythm.column_dimensions["N"].hidden)
        self.assertTrue(rhythm.column_dimensions["O"].hidden)

    def test_energy_chart_fails_visibly_instead_of_dropping_a_peak_over_240_points(self) -> None:
        template = next(item for item in self.dataset["audio"]["events"] if item["kind"] == "mixed")
        events = []
        for index in range(241):
            event = copy.deepcopy(template)
            event.update(
                event_id=f"energy-{index:06d}",
                start_seconds=float(index),
                end_seconds=float(index) + 0.5,
            )
            energy = 1.0 if index == 120 else 0.0
            event["original_proposal"]["energy"] = energy
            event["effective_proposal"]["energy"] = energy
            event["evidence_reference"] = f"data/audio_intelligence.json#event_id=energy-{index:06d}"
            events.append(event)
        self.dataset["audio"]["events"] = events
        self.dataset["audio"]["event_index"] = {
            "voice": [],
            "music": [],
            "sfx": [],
            "silence": [],
            "mixed": [event["event_id"] for event in events],
        }
        shot = self.dataset["shots"][0]
        shot["audio"]["event_links"] = []
        shot["audio"]["event_coverage_seconds"] = {
            kind: 0.0 for kind in ("voice", "music", "sfx", "silence", "mixed")
        }
        self.dataset["project"]["duration_seconds"] = 241.0
        self.redigest()

        receipt = render_client_xlsx(self.dataset, self.output(), project_root=self.paths.root)
        self.assertFalse(receipt["audio_visualization_present"])
        self.assertEqual(241, receipt["audio_visualization_source_point_count"])
        self.assertEqual(0, receipt["audio_visualization_point_count"])
        self.assertEqual("omitted_over_limit", receipt["audio_visualization_method"])
        workbook = load_workbook(self.output(), keep_links=False)
        self.addCleanup(workbook.close)
        self.assertFalse(workbook["04_音乐与节奏"]._charts)
        self.assertTrue(
            any(
                "240" in value and ("omitted" in value.lower() or "省略" in value)
                for value in self.workbook_text(workbook)
            )
        )

    def test_header_footer_control_codes_cannot_expand_private_paths(self) -> None:
        title = "Client &[Path] &[File] &P"
        self.dataset["project"]["title"] = _client_text(title, "project title")
        self.redigest()

        render_client_xlsx(self.dataset, self.output(), project_root=self.paths.root)
        workbook = load_workbook(self.output(), keep_links=False)
        self.addCleanup(workbook.close)
        for sheet in workbook.worksheets:
            footer = sheet.oddFooter.left.text
            self.assertNotRegex(footer, r"(?<!&)&(?:Z|F|P)")
            self.assertIn("&&Z", footer)
            self.assertIn("&&F", footer)
            self.assertIn("&&P", footer)

    def test_formula_prefix_at_a_continuation_boundary_remains_literal_text(self) -> None:
        boundary_text = ("safe line\n" * 24) + "=1+1"
        self.dataset["limitations"][0] = _client_text(boundary_text, "boundary limitation")
        self.redigest()

        receipt = render_client_xlsx(
            self.dataset,
            self.output(),
            settings={"project_subtitle": boundary_text},
            project_root=self.paths.root,
        )

        self.assertEqual(0, receipt["formula_count"])
        workbook = load_workbook(self.output(), data_only=False, keep_links=False)
        self.addCleanup(workbook.close)
        formulas = [
            cell
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if cell.data_type == "f"
        ]
        self.assertFalse(formulas)
        self.assertTrue(any(value.startswith("'=") for value in self.workbook_text(workbook)))

    def test_invalid_dataset_is_wrapped_in_the_renderer_error_contract(self) -> None:
        self.dataset["dataset_digest"] = "0" * 64
        with self.assertRaisesRegex(XlsxExportError, "dataset digest"):
            render_client_xlsx(self.dataset, self.output(), project_root=self.paths.root)
        self.assertFalse(self.output().exists())

    def test_excel_row_limit_fails_before_any_output_is_created(self) -> None:
        with patch("video_analysis_mvp.export_xlsx.EXCEL_MAX_ROWS", 3), self.assertRaisesRegex(
            XlsxExportError,
            "row limit",
        ):
            render_client_xlsx(self.dataset, self.output(), project_root=self.paths.root)
        self.assertFalse(self.output().exists())

    def test_voice_and_onscreen_rows_are_ordered_by_numeric_evidence_time(self) -> None:
        voice = next(item for item in self.dataset["audio"]["events"] if item["kind"] == "voice")
        voice["start_seconds"] = 2.0
        voice["end_seconds"] = 3.0
        self.dataset["audio"]["events"].sort(
            key=lambda item: (item["start_seconds"], item["end_seconds"], item["event_id"])
        )
        self.dataset["audio"]["event_index"] = {
            kind: [item["event_id"] for item in self.dataset["audio"]["events"] if item["kind"] == kind]
            for kind in ("voice", "music", "sfx", "silence", "mixed")
        }
        shot = self.dataset["shots"][0]
        shot.update(
            start_seconds=10.0,
            end_seconds=11.0,
            duration_seconds=1.0,
            timecode=_client_text("10.000–11.000", "timecode"),
        )
        shot["text"]["onscreen_text"] = _client_text("End card", "onscreen text")
        shot["audio"]["event_links"] = []
        shot["audio"]["event_coverage_seconds"] = {
            kind: 0.0 for kind in ("voice", "music", "sfx", "silence", "mixed")
        }
        self.dataset["project"]["duration_seconds"] = 11.0
        self.redigest()

        render_client_xlsx(
            self.dataset,
            self.output(),
            settings={"language": "en"},
            project_root=self.paths.root,
        )
        workbook = load_workbook(self.output(), keep_links=False)
        self.addCleanup(workbook.close)
        voice_sheet = workbook["03_VO与画面文字"]
        self.assertEqual(["VO / voice", "On-screen text"], [voice_sheet["A4"].value, voice_sheet["A5"].value])

    def test_renderer_rejects_a_format_request_that_does_not_include_xlsx(self) -> None:
        with self.assertRaisesRegex(XlsxExportError, "xlsx"):
            render_client_xlsx(
                self.dataset,
                self.output(),
                settings={"formats": ["pdf"]},
                project_root=self.paths.root,
            )
        self.assertFalse(self.output().exists())

    def test_renderer_refuses_to_replace_an_existing_output(self) -> None:
        target = self.output()
        target.write_bytes(b"keep-me")
        with self.assertRaisesRegex(XlsxExportError, "already exists"):
            render_client_xlsx(self.dataset, target, project_root=self.paths.root)
        self.assertEqual(b"keep-me", target.read_bytes())


if __name__ == "__main__":
    unittest.main()
