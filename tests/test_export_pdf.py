from __future__ import annotations

import copy
import importlib.util
import os
import re
import tempfile
import unicodedata
import unittest
from pathlib import Path
from unittest.mock import patch

from tests.test_audio_review import audio_review_fixture
from tests.test_evidence_handoff import PNG_1X1
from video_analysis_mvp import export_pdf as pdf_module
from video_analysis_mvp.client_export_dataset import (
    _canonical_digest,
    _client_text,
    build_client_export_dataset,
)
from video_analysis_mvp.export_pdf import (
    PdfExportError,
    _fonts_embedded,
    render_client_html,
    render_client_pdf,
)
from video_analysis_mvp.export_xlsx import render_client_xlsx
from video_analysis_mvp.image_evidence import inspect_image_bytes
from video_analysis_mvp.utils import ProcessCancelledError, ToolError

NODE = Path(os.environ.get("VEW_PDF_NODE", "__vew_pdf_node_unavailable__"))
NODE_MODULES = Path(
    os.environ.get("VEW_PDF_NODE_MODULES", "__vew_pdf_modules_unavailable__")
)
CHROME = Path(os.environ.get("VEW_PDF_BROWSER", "__vew_pdf_browser_unavailable__"))
FONT = Path(os.environ.get("VEW_PDF_FONT", "__vew_pdf_font_unavailable__"))
NON_CJK_FONT = Path(os.environ.get("VEW_PDF_NON_CJK_FONT", "/System/Library/Fonts/Symbol.ttf"))
PDF_RUNTIME = NODE.is_file() and (NODE_MODULES / "playwright").is_dir() and CHROME.is_file() and FONT.is_file()
PYPDF_AVAILABLE = importlib.util.find_spec("pypdf") is not None
OPENPYXL_AVAILABLE = importlib.util.find_spec("openpyxl") is not None
if os.environ.get("VEW_REQUIRE_PDF_RUNTIME") == "1" and not (PDF_RUNTIME and PYPDF_AVAILABLE):
    raise RuntimeError("VEW_REQUIRE_PDF_RUNTIME=1 but the real Playwright PDF runtime is unavailable")


class PdfExportTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory(prefix="vew-pdf-")
        self.addCleanup(self.temp.cleanup)
        self.paths = audio_review_fixture(Path(self.temp.name))
        self.dataset = build_client_export_dataset(self.paths)

    def redigest(self) -> None:
        base = {key: value for key, value in self.dataset.items() if key not in {"dataset_id", "dataset_digest"}}
        digest = _canonical_digest(base)
        self.dataset["dataset_id"] = self.dataset["dataset_digest"] = digest

    def html_path(self) -> Path:
        return self.paths.root / "client-breakdown.html"

    def pdf_path(self) -> Path:
        return self.paths.root / "client-breakdown.pdf"

    def settings(self) -> dict[str, object]:
        return {"language": "bilingual", "project_subtitle": "Evidence review"}

    def test_html_is_complete_escaped_self_contained_and_print_ready(self) -> None:
        dangerous = '<script>alert(1)</script> & =SUM(1+1)'
        shot = self.dataset["shots"][0]
        shot["text"]["content_summary"] = _client_text(dangerous, "dangerous summary")
        shot["text"]["content_summary_zh"] = _client_text("中文镜头说明", "Chinese summary")
        shot["text"]["onscreen_text"] = _client_text("ONSCREEN-MARKER", "onscreen")
        shot["camera"]["motion"] = _client_text("CAMERA-MARKER", "camera")
        shot["text"]["transition_out"] = _client_text("TRANSITION-MARKER", "transition")
        self.redigest()

        receipt = render_client_html(
            self.dataset,
            self.html_path(),
            settings=self.settings(),
            project_root=self.paths.root,
            available_fonts=["Noto Sans CJK SC"],
        )

        rendered = self.html_path().read_text(encoding="utf-8")
        self.assertEqual("html-render-receipt/v1", receipt["schema_id"])
        self.assertEqual(self.dataset["dataset_digest"], receipt["dataset_digest"])
        self.assertEqual(7, receipt["section_count"])
        self.assertEqual(1, receipt["shot_count"])
        self.assertEqual(len(self.dataset["audio"]["events"]), receipt["audio_event_count"])
        self.assertIn("@page { size: A4 landscape", rendered)
        self.assertIn("font-size: 8.5pt", rendered)
        self.assertIn("&lt;script&gt;alert(1)&lt;/script&gt; &amp; =SUM(1+1)", rendered)
        self.assertNotIn("<script", rendered.lower())
        self.assertNotIn("file://", rendered)
        self.assertNotIn(str(self.paths.root), rendered)
        self.assertNotIn("http://", rendered)
        self.assertNotIn("https://", rendered)
        self.assertIn('<html lang="zh-Hans">', rendered)
        self.assertIn("ONSCREEN-MARKER", rendered)
        self.assertIn("CAMERA-MARKER", rendered)
        self.assertIn("TRANSITION-MARKER", rendered)
        self.assertLess(rendered.index("shot_0001</td>"), rendered.index("voice-1</td>"))
        for section in ("cover", "overview", "narrative-audio", "storyboard", "voice-text", "audio-rhythm", "evidence"):
            self.assertIn(f'id="{section}"', rendered)
        self.assertFalse(list(self.paths.root.rglob("*.pdf")))

    def test_html_preview_does_not_require_a_pdf_font_identity_claim(self) -> None:
        receipt = render_client_html(
            self.dataset,
            self.html_path(),
            settings=self.settings(),
            project_root=self.paths.root,
            available_fonts=[],
        )
        self.assertEqual(["html"], receipt["settings"]["formats"])

    @unittest.skipUnless(OPENPYXL_AVAILABLE, "spreadsheet runtime is unavailable")
    def test_html_and_xlsx_cover_the_same_client_shot_fields(self) -> None:
        from openpyxl import load_workbook

        markers = {
            "SUMMARY-MARKER": ("text", "content_summary_zh"),
            "DIALOGUE-MARKER": ("text", "dialogue"),
            "ONSCREEN-MARKER": ("text", "onscreen_text"),
            "MUSIC-MARKER": ("text", "music_state"),
            "SFX-MARKER": ("text", "sound_design"),
            "RHYTHM-MARKER": ("text", "rhythm_notes"),
            "TRANSITION-MARKER": ("text", "transition_out"),
            "CAMERA-MARKER": ("camera", "motion"),
        }
        shot = self.dataset["shots"][0]
        for marker, (group, field) in markers.items():
            shot[group][field] = _client_text(marker, field)
        self.redigest()
        render_client_html(
            self.dataset,
            self.html_path(),
            settings=self.settings(),
            project_root=self.paths.root,
            available_fonts=[],
        )
        xlsx = self.paths.root / "client.xlsx"
        render_client_xlsx(self.dataset, xlsx, settings=self.settings(), project_root=self.paths.root)
        workbook = load_workbook(xlsx, keep_links=False)
        self.addCleanup(workbook.close)
        html_text = self.html_path().read_text(encoding="utf-8")
        xlsx_text = "\n".join(
            cell.value
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str)
        )
        for marker in markers:
            self.assertIn(marker, html_text)
            self.assertIn(marker, xlsx_text)

    def test_html_embeds_bound_frames_and_rejects_changed_bytes(self) -> None:
        frame = self.paths.keyframes / "frame.png"
        frame.write_bytes(PNG_1X1)
        self.dataset["shots"][0]["frame"] = {
            "path": "assets/keyframes/frame.png",
            "present": True,
            **inspect_image_bytes(PNG_1X1).receipt_fields(),
            "failure": _client_text("", "frame failure"),
        }
        self.redigest()

        receipt = render_client_html(
            self.dataset,
            self.html_path(),
            settings={**self.settings(), "logo_path": "assets/keyframes/frame.png"},
            project_root=self.paths.root,
            available_fonts=["Noto Sans CJK SC"],
        )
        self.assertEqual(2, receipt["embedded_image_count"])
        self.assertEqual(2, self.html_path().read_text(encoding="utf-8").count("data:image/png;base64,"))

        frame.write_bytes(PNG_1X1 + b"changed")
        with self.assertRaisesRegex(PdfExportError, "frame evidence"):
            render_client_html(
                self.dataset,
                self.paths.root / "changed.html",
                settings=self.settings(),
                project_root=self.paths.root,
                available_fonts=["Noto Sans CJK SC"],
            )
        self.assertFalse((self.paths.root / "changed.html").exists())

    def test_pdf_requires_an_explicit_existing_renderer_and_never_downloads(self) -> None:
        with self.assertRaisesRegex(PdfExportError, "Playwright"):
            render_client_pdf(
                self.dataset,
                self.pdf_path(),
                settings=self.settings(),
                project_root=self.paths.root,
                available_fonts=["Noto Sans CJK SC"],
                node_executable=self.paths.root / "missing-node",
                node_modules_path=self.paths.root / "missing-modules",
                browser_executable=self.paths.root / "missing-browser",
            )
        self.assertFalse(self.pdf_path().exists())

    def test_renderer_failure_diagnostic_is_bounded_and_path_free(self) -> None:
        formatter = getattr(pdf_module, "_renderer_failure_detail", None)
        self.assertIsNotNone(formatter)
        private_path = "/Users/example/private/browser"
        detail = formatter(
            "Command failed\n"
            f"{private_path}\n"
            "RendererDiagnostic:launch-browser:browser-launch-failed:Error"
        )

        self.assertEqual(
            " [stage=launch-browser, code=browser-launch-failed, name=Error]",
            detail,
        )
        self.assertNotIn(private_path, detail)
        self.assertEqual(
            "",
            formatter(
                "/tmp/RendererDiagnostic:launch-browser:browser-launch-failed:Error/output"
            ),
        )
        self.assertEqual(
            "",
            formatter(
                "RendererDiagnostic:launch-browser:browser-launch-failed:"
                + "E" * 65
            ),
        )
        self.assertEqual(
            "",
            formatter(
                "RendererDiagnostic:launch-browser:browser-launch-failed:Error trailing"
            ),
        )
        self.assertIn(
            "RendererDiagnostic:${stage}:${code}:${name}",
            pdf_module.DRIVER_PATH.read_text(encoding="utf-8"),
        )

    @unittest.skipUnless(PDF_RUNTIME and PYPDF_AVAILABLE, "dedicated PDF runtime is unavailable")
    def test_pdf_rejects_a_fake_font_without_publishing_an_output(self) -> None:
        fake_font = self.paths.root / "fake-font.otf"
        fake_font.write_bytes(b"OTTO" + (b"not-a-font" * 32))

        with self.assertRaisesRegex(PdfExportError, "cannot be decoded"):
            render_client_pdf(
                self.dataset,
                self.pdf_path(),
                settings=self.settings(),
                project_root=self.paths.root,
                available_fonts=["Noto Sans CJK SC"],
                node_executable=NODE,
                node_modules_path=NODE_MODULES,
                browser_executable=CHROME,
                font_path=fake_font,
            )
        self.assertFalse(self.pdf_path().exists())

    @unittest.skipUnless(PDF_RUNTIME and PYPDF_AVAILABLE, "dedicated PDF runtime is unavailable")
    def test_pdf_cancellation_uses_the_bounded_process_group_path(self) -> None:
        with (
            patch(
                "video_analysis_mvp.export_pdf.run_command",
                side_effect=ProcessCancelledError("cancelled"),
            ) as runner,
            self.assertRaisesRegex(PdfExportError, "cancelled"),
        ):
            render_client_pdf(
                self.dataset,
                self.pdf_path(),
                settings=self.settings(),
                project_root=self.paths.root,
                available_fonts=["Noto Sans CJK SC"],
                node_executable=NODE,
                node_modules_path=NODE_MODULES,
                browser_executable=CHROME,
                font_path=FONT,
                cancelled=lambda: True,
            )

        self.assertEqual(2 * 1024 * 1024, runner.call_args.kwargs["max_output_bytes"])
        self.assertFalse(self.pdf_path().exists())

    @unittest.skipUnless(PDF_RUNTIME and PYPDF_AVAILABLE, "dedicated PDF runtime is unavailable")
    def test_pdf_browser_runtime_temp_is_private_and_outside_output_tree(self) -> None:
        observed: dict[str, object] = {}

        def capture_and_fail(*_args, **kwargs):
            environment = kwargs["environment"]
            runtime_temp = Path(environment["TMPDIR"])
            observed.update(
                {
                    "home": environment["HOME"],
                    "tmpdir": environment["TMPDIR"],
                    "mode": runtime_temp.stat().st_mode & 0o777,
                }
            )
            raise ToolError(
                "Command failed\n"
                "RendererDiagnostic:launch-browser:browser-launch-failed:Error"
            )

        with (
            patch("video_analysis_mvp.export_pdf.run_command", side_effect=capture_and_fail),
            self.assertRaisesRegex(PdfExportError, "stage=launch-browser"),
        ):
            render_client_pdf(
                self.dataset,
                self.pdf_path(),
                settings=self.settings(),
                project_root=self.paths.root,
                available_fonts=["Noto Sans CJK SC"],
                node_executable=NODE,
                node_modules_path=NODE_MODULES,
                browser_executable=CHROME,
                font_path=FONT,
            )

        runtime_temp = Path(str(observed["tmpdir"]))
        self.assertEqual(observed["home"], observed["tmpdir"])
        self.assertEqual(0o700, observed["mode"])
        self.assertFalse(runtime_temp.is_relative_to(self.pdf_path().parent))
        self.assertFalse(runtime_temp.exists())

    @unittest.skipUnless(PDF_RUNTIME and PYPDF_AVAILABLE, "dedicated PDF runtime is unavailable")
    def test_pdf_does_not_claim_cancelled_when_group_cleanup_is_unverified(self) -> None:
        with (
            patch(
                "video_analysis_mvp.export_pdf.run_command",
                side_effect=ProcessCancelledError(
                    "cancelled", cleanup_verified=False
                ),
            ),
            self.assertRaisesRegex(PdfExportError, "cleanup could not be verified") as caught,
        ):
            render_client_pdf(
                self.dataset,
                self.pdf_path(),
                settings=self.settings(),
                project_root=self.paths.root,
                available_fonts=["Noto Sans CJK SC"],
                node_executable=NODE,
                node_modules_path=NODE_MODULES,
                browser_executable=CHROME,
                font_path=FONT,
                cancelled=lambda: True,
            )

        self.assertIs(caught.exception.process_group_cleanup_verified, False)
        self.assertFalse(self.pdf_path().exists())

    @unittest.skipUnless(
        PDF_RUNTIME and PYPDF_AVAILABLE and NON_CJK_FONT.is_file(),
        "valid non-CJK font fixture is unavailable",
    )
    def test_pdf_rejects_a_valid_font_without_required_cjk_glyphs(self) -> None:
        with self.assertRaisesRegex(PdfExportError, "lacks required CJK glyphs"):
            render_client_pdf(
                self.dataset,
                self.pdf_path(),
                settings=self.settings(),
                project_root=self.paths.root,
                available_fonts=["Noto Sans CJK SC"],
                node_executable=NODE,
                node_modules_path=NODE_MODULES,
                browser_executable=CHROME,
                font_path=NON_CJK_FONT,
            )
        self.assertFalse(self.pdf_path().exists())

    @unittest.skipUnless(PDF_RUNTIME and PYPDF_AVAILABLE, "dedicated PDF runtime is unavailable")
    def test_real_pdf_has_a4_landscape_text_images_metadata_and_receipt(self) -> None:
        from pypdf import PdfReader

        frame = self.paths.keyframes / "frame.png"
        frame.write_bytes(PNG_1X1)
        self.dataset["shots"][0]["frame"] = {
            "path": "assets/keyframes/frame.png",
            "present": True,
            **inspect_image_bytes(PNG_1X1).receipt_fields(),
            "failure": _client_text("", "frame failure"),
        }
        self.dataset["shots"][0]["text"]["content_summary_zh"] = _client_text("客户可审阅的镜头文本", "summary")
        self.dataset["shots"][0]["text"]["onscreen_text"] = _client_text("CAPTION " + ("caption text " * 24), "caption")
        self.redigest()

        receipt = render_client_pdf(
            self.dataset,
            self.pdf_path(),
            settings=self.settings(),
            project_root=self.paths.root,
            available_fonts=["Noto Sans CJK SC"],
            node_executable=NODE,
            node_modules_path=NODE_MODULES,
            browser_executable=CHROME,
            font_path=FONT,
        )

        reader = PdfReader(self.pdf_path())
        self.assertGreaterEqual(len(reader.pages), 7)
        first = reader.pages[0]
        self.assertGreater(float(first.mediabox.width), float(first.mediabox.height))
        self.assertAlmostEqual(841.9, float(first.mediabox.width), delta=2)
        self.assertAlmostEqual(595.3, float(first.mediabox.height), delta=2)
        text = "\n".join(page.extract_text() or "" for page in reader.pages)
        self.assertIn("客户可审阅的镜头文本", unicodedata.normalize("NFKC", text))
        self.assertIn("Original VO", text)
        self.assertIn(self.dataset["dataset_digest"], text)
        self.assertEqual("Video Evidence Workbench", reader.metadata.author)
        self.assertEqual("pdf-render-receipt/v1", receipt["schema_id"])
        self.assertEqual(len(reader.pages), receipt["page_count"])
        self.assertEqual(1, receipt["embedded_image_count"])
        self.assertTrue(receipt["searchable_text"])
        self.assertTrue(receipt["embedded_fonts"])
        self.assertEqual("A4-landscape", receipt["page_size"])
        self.assertTrue(receipt["font_plan"]["identity_verified"])
        self.assertRegex(receipt["font_plan"]["font_sha256"], r"^[0-9a-f]{64}$")
        self.assertGreater(receipt["font_plan"]["verified_cjk_glyphs"], 0)
        self.assertIn("Noto Sans CJK", receipt["font_plan"]["resolved_font_family"])
        self.assertTrue(receipt["renderer"]["font_loaded"])

    @unittest.skipUnless(PDF_RUNTIME and PYPDF_AVAILABLE, "dedicated PDF runtime is unavailable")
    def test_long_story_text_flows_to_additional_pages_without_losing_the_end(self) -> None:
        from pypdf import PdfReader

        long_text = "BEGIN-LONG-TEXT " + ("长文本证据 " * 4_000) + " END-LONG-TEXT"
        self.dataset["shots"][0]["text"]["content_summary_zh"] = _client_text(long_text, "long PDF text")
        self.redigest()

        receipt = render_client_pdf(
            self.dataset,
            self.pdf_path(),
            settings=self.settings(),
            project_root=self.paths.root,
            available_fonts=["Noto Sans CJK SC"],
            node_executable=NODE,
            node_modules_path=NODE_MODULES,
            browser_executable=CHROME,
            font_path=FONT,
        )
        text = unicodedata.normalize(
            "NFKC",
            "\n".join(page.extract_text() or "" for page in PdfReader(self.pdf_path()).pages),
        )
        compact = re.sub(r"\s+", "", text)
        self.assertIn("BEGIN-LONG-TEXT", compact)
        self.assertIn("END-LONG-TEXT", compact)
        self.assertGreater(compact.count("#0001"), 1)
        self.assertGreater(receipt["page_count"], 8)

    def test_font_embedding_requires_every_font_to_prove_embedded_bytes(self) -> None:
        class Ref:
            def __init__(self, value):
                self.value = value

            def get_object(self):
                return self.value

        embedded = Ref({"/Subtype": "/Type1", "/FontDescriptor": Ref({"/FontFile2": object()})})
        missing = Ref({"/Subtype": "/Type1"})
        type3_good = Ref({"/Subtype": "/Type3", "/CharProcs": {"A": object()}})
        type3_bad = Ref({"/Subtype": "/Type3"})

        class Reader:
            def __init__(self, fonts):
                self.pages = [{"/Resources": {"/Font": fonts}}]

        self.assertTrue(_fonts_embedded(Reader({"A": embedded, "B": type3_good})))
        self.assertFalse(_fonts_embedded(Reader({"A": embedded, "B": missing})))
        self.assertFalse(_fonts_embedded(Reader({"A": embedded, "B": type3_bad})))

    def test_two_hundred_five_shots_are_all_present_in_html_without_truncation(self) -> None:
        source = self.dataset["shots"][0]
        shots = []
        zero_coverage = {kind: 0.0 for kind in ("voice", "music", "sfx", "silence", "mixed")}
        for index in range(1, 206):
            shot = copy.deepcopy(source)
            shot.update(
                shot_id=f"shot_{index:04d}",
                shot_no=index,
                start_seconds=float(index - 1),
                end_seconds=float(index),
                duration_seconds=1.0,
                timecode=_client_text(f"{index - 1:04d}-{index:04d}", "timecode"),
                evidence_reference=f"data/shots.json#shot_id=shot_{index:04d}",
            )
            shot["frame"] = {
                "path": f"assets/keyframes/frame-{index:04d}.jpg",
                "present": False,
                "sha256": None,
                "size_bytes": None,
                "media_type": None,
                "width": None,
                "height": None,
                "failure": _client_text("Primary frame unavailable", "frame failure"),
            }
            shot["audio"]["event_links"] = []
            shot["audio"]["event_coverage_seconds"] = dict(zero_coverage)
            shots.append(shot)
        self.dataset["shots"] = shots
        self.dataset["project"]["duration_seconds"] = 205.0
        self.redigest()

        receipt = render_client_html(
            self.dataset,
            self.html_path(),
            settings=self.settings(),
            project_root=self.paths.root,
            available_fonts=["Noto Sans CJK SC"],
        )
        rendered = self.html_path().read_text(encoding="utf-8")
        self.assertEqual(205, receipt["shot_count"])
        self.assertEqual(205, rendered.count('class="shot-card'))
        self.assertIn("shot_0205", rendered)
        if PDF_RUNTIME and PYPDF_AVAILABLE:
            receipt = render_client_pdf(
                self.dataset,
                self.pdf_path(),
                settings=self.settings(),
                project_root=self.paths.root,
                available_fonts=["Noto Sans CJK SC"],
                node_executable=NODE,
                node_modules_path=NODE_MODULES,
                browser_executable=CHROME,
                font_path=FONT,
            )
            self.assertLessEqual(receipt["page_count"], 65)


if __name__ == "__main__":
    unittest.main()
